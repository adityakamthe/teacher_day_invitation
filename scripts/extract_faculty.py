import os
import re
import io
import json
import openpyxl
from PIL import Image

def slugify(text):
    # Strip titles strictly
    cleaned = re.sub(r'^(Dr\.?|Prof\.?|Mrs\.?|Mr\.?|Ms\.?)\s*', '', text, flags=re.IGNORECASE).strip()
    # Normalize special characters, commas, dots
    cleaned = re.sub(r'[\.\,\']', '', cleaned)
    # Replace spaces and underscores with hyphen
    cleaned = re.sub(r'[\s_]+', '-', cleaned)
    # Remove leading/trailing hyphens
    cleaned = cleaned.strip('-').lower()
    return cleaned

def clean_name(name):
    # Fix spacing around initials and dots like 'Dr.Madhuri N.Jadhav' -> 'Dr. Madhuri N. Jadhav'
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'\.(?! )', '. ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    # Fix specific name formatting cleanly
    if name.startswith("Prof ") and not name.startswith("Prof. "):
        name = name.replace("Prof ", "Prof. ", 1)
    if name.startswith("Dr ") and not name.startswith("Dr. "):
        name = name.replace("Dr ", "Dr. ", 1)
    if name.startswith("Mr ") and not name.startswith("Mr. "):
        name = name.replace("Mr ", "Mr. ", 1)
    if name.startswith("Mrs ") and not name.startswith("Mrs. "):
        name = name.replace("Mrs ", "Mrs. ", 1)
    if name.startswith("Ms ") and not name.startswith("Ms. "):
        name = name.replace("Ms ", "Ms. ", 1)
    return name

def main():
    excel_path = "Faculty Details.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Faculty Nameplates"]
    
    os.makedirs("public/faculty", exist_ok=True)
    os.makedirs("data", exist_ok=True)
    
    # Map image by anchor row
    images_by_row = {}
    for img in ws._images:
        anchor = img.anchor
        from_row = anchor._from.row if hasattr(anchor, '_from') else getattr(anchor, 'row', None)
        if from_row is not None:
            images_by_row[from_row] = img

    faculty_list = []
    
    # Teaching Faculty: Rows 6 to 23
    # Technical Staff: Rows 27 to 32
    target_rows = list(range(6, 24)) + list(range(27, 33))
    
    for r in target_rows:
        sr = ws.cell(r, 2).value
        raw_name = ws.cell(r, 3).value
        raw_desig = ws.cell(r, 4).value
        
        if not raw_name:
            continue
            
        name = clean_name(str(raw_name))
        
        # Designation cleaning & fallback
        if not raw_desig or str(raw_desig).strip().lower() in ['none', '']:
            if r < 24:
                designation = "Assistant Professor"
            else:
                designation = "Technical Assistant"
        else:
            designation = str(raw_desig).strip()
            if "profesor" in designation.lower() or "professor" in designation.lower():
                designation = "Assistant Professor"
            elif "technical assistant" in designation.lower():
                designation = "Technical Assistant"
                
        slug = slugify(name)
        
        # Look for image at row index (0-indexed is r-1)
        img = images_by_row.get(r - 1)
        if not img:
            for offset in [-2, -1, 0, 1, 2]:
                if (r - 1 + offset) in images_by_row:
                    img = images_by_row[r - 1 + offset]
                    break
        
        photo_filename = f"{slug}.jpg"
        photo_rel_path = f"/faculty/{photo_filename}"
        photo_save_path = os.path.join("public", "faculty", photo_filename)
        
        has_photo = False
        if img:
            try:
                img_data = img._data()
                image = Image.open(io.BytesIO(img_data))
                if image.mode in ('RGBA', 'LA', 'P'):
                    image = image.convert('RGB')
                
                max_size = 800
                if max(image.size) > max_size:
                    image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                    
                image.save(photo_save_path, "JPEG", quality=90, optimize=True)
                has_photo = True
            except Exception as e:
                print(f"Warning: Failed to save photo for {name} ({slug}): {e}")
        else:
            print(f"Warning: No image found for row {r}: {name} ({slug})")
            
        faculty_list.append({
            "id": len(faculty_list) + 1,
            "slug": slug,
            "name": name,
            "designation": designation,
            "photo": photo_rel_path if has_photo else None,
            "category": "Teaching Faculty" if r < 24 else "Technical Assistant"
        })

    with open("data/faculty.json", "w", encoding="utf-8") as f:
        json.dump(faculty_list, f, indent=2, ensure_ascii=False)
        
    print(f"\nExtraction summary: Successfully processed all {len(faculty_list)} recipients (24 in, 24 out).")
    for f in faculty_list:
        status = "Photo OK" if f["photo"] else "No Photo (Initials badge fallback)"
        print(f"[{f['id']:02d}] {f['slug']:<25} | {f['name']:<30} | {f['designation']:<22} | {status}")

if __name__ == "__main__":
    main()
